import pandas as pd
pd.options.mode.chained_assignment = None  # default='warn'people

from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
from tqdm import tqdm
import re
import os
from datetime import datetime, timedelta

def analyze_redundant_policies(df, vendor, file_name):
    logging.info("Redundant Policies Analysis Started")
    try:
        logging.info('Parsing firewall policies')
        df_filtered = df[(df['Enable'] == 'Y') & (df['Action'] == 'Allow')]

        columns_to_check = ['Enable', 'Action', 'Source', 'User', 'Destination', 'Service', 'Application']
        if 'Vsys' in df.columns:
            columns_to_check.append('Vsys')
        
        if vendor == 'paloalto':
            df_filtered['Service'] = df_filtered['Service'].str.replace('_','-')
            columns_to_check.append('Category')
        
        df_check = df_filtered[columns_to_check]

        def normalize_policy(policy_series):
            normalized_policy = policy_series.apply(lambda x: ','.join(sorted(x.split(','))) if isinstance(x, str) else x)
            return tuple(normalized_policy)
        
        policy_map = defaultdict(list)

        results_list = []
        current_no = 1

        logging.info('Checking for redundant policies')

        for i in tqdm(range(len(df_filtered)), desc='Checking Policies'):
            try:
                current_policy = normalize_policy(df_check.iloc[i])
                if current_policy in policy_map:
                    row = df_filtered.iloc[i].to_dict()
                    row.update({'No': policy_map[current_policy], 'Type': 'Lower'})
                    results_list.append(row)
                else:
                    policy_map[current_policy] = current_no
                    row = df_filtered.iloc[i].to_dict()
                    row.update({'No': current_no, 'Type': 'Upper'})
                    results_list.append(row)
                    current_no += 1
            except Exception as e:
                logging.error(f'Error in checking policy at index {i}: {e}')
                continue
        
        logging.info('Converting results to DataFrame')
        results = pd.DataFrame(results_list)

        logging.info('Ensuring each No group contains both Upper and Lower.')
        def ensure_upper_and_lower(df):
            valid_no_groups = []
            grouped = df.groupby('No')
            for name, group in grouped:
                if 'Upper' in group['Type'].values and 'Lower' in group['Type'].values:
                    valid_no_groups.append(group)
            return pd.concat(valid_no_groups).reset_index(drop=True)
    
        duplicated_results = ensure_upper_and_lower(results)

        duplicated_results['No'] = duplicated_results.groupby('No').ngroup() + 1

        columns_order = ['No', 'Type'] + [col for col in df.columns]
        duplicated_results = duplicated_results[columns_order]

        duplicated_results = duplicated_results.sort_values(by=['No', 'Type'], ascending=[True, False])

        # style
        upper_fill = PatternFill(start_color="daeef3", end_color="daeef3", fill_type="solid")
        lower_fill = PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type="solid")
        header_fill = PatternFill(start_color="00b0f0", end_color="00b0f0", fill_type="solid")
        header_font = Font(bold=True, color='FFFFFF')

        logging.info("saving results to excel")
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            if 'vsys' in df.columns:
                for vsys, vsys_df in duplicated_results.groupby('vsys'):
                    vsys_df.to_excel(writer, index=False, sheet_name=f'Analysis_{vsys}')
                    worksheet = writer.sheets[f'Analysis_{vsys}']

                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                        for cell in row:
                            if row[1].value == 'Upper':
                                cell.fill = upper_fill
                            elif row[1].value == 'Lower':
                                cell.fill = lower_fill
            else:
                duplicated_results.to_excel(writer, index=False, sheet_name='Analysis')
                worksheet = writer.sheets['Analysis']

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        if row[1].value == 'Upper':
                            cell.fill = upper_fill
                        elif row[1].value == 'Lower':
                            cell.fill = lower_fill
        
        logging.info(f"Results have been saved to {file_name}")
    except Exception as e:
        logging.error(f"Error in analyzing redundant policies: {e}")
    
    logging.info("Redundant Policies Analysis Completed")

def compare_and_save_firewall_policies(df_before, df_after, output_filename='firewall_policy_changes.xlsx'):
    def compare_firewall_policies(df_before, df_after):
        # Merge the two dataframes on 'Rule Name' column to compare row by row
        df_merged = df_before.merge(df_after, on='Rule Name', how='outer', suffixes=('_before', '_after'), indicator=True)

        # Find added, removed, and changed rows
        added = df_merged[df_merged['_merge'] == 'right_only']
        removed = df_merged[df_merged['_merge'] == 'left_only']

        # Ensure 'Rule Name' is always included in added and removed rows
        added_col = ['Rule Name'] + [col for col in added.columns if col.endswith('_after')]
        removed_col = ['Rule Name'] + [col for col in removed.columns if col.endswith('_before')]

        added = added[added_col].rename(columns=lambda x: x.replace('_after', ''))
        removed = removed[removed_col].rename(columns=lambda x: x.replace('_before', ''))

        # Identify changed rows, ignoring changes in 'Seq'
        common_cols = [col for col in df_before.columns if col not in ['Rule Name', 'Seq']]
        change_conditions = [df_merged[f'{col}_before'] != df_merged[f'{col}_after'] for col in common_cols]
        changed = df_merged[(df_merged['_merge'] == 'both') & pd.concat(change_conditions, axis=1).any(axis=1)]

        # create a dataframe to store only the changes
        changes_list = []
        for idx, row in changed.iterrows():
            changes = {'Rule Name': row['Rule Name']}
            for col in common_cols:
                if row[f'{col}_before'] != row[f'{col}_after']:
                    changes[f'{col}_before'] = row[f'{col}_before']
                    changes[f'{col}_after'] = row[f'{col}_after']
            changes_list.append(changes)
            changed_df = pd.DataFrame(changes_list)

            return added, removed, changed_df
    
    def display_and_save_results(added, removed, changed, output_filename):
        added_count = len(added)
        removed_count = len(removed)
        changed_count = len(changed)

        logging.info(f"Added: {added_count} rows")
        logging.info(f"Removed: {removed_count} rows")
        logging.info(f"Changed: {changed_count} rows")

        logging.info(f"Saving results to {output_filename}")
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            summary_data = {
                'Category': ['Added', 'Removed', 'Changed'],
                'Count': [added_count, removed_count, changed_count]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, index=False, sheet_name='Summary')

            added.to_excel(writer, index=False, sheet_name='Added')
            removed.to_excel(writer, index=False, sheet_name='Removed')
            changed.to_excel(writer, index=False, sheet_name='Changed')
        
        logging.info(f"Results have been saved to {output_filename}")
    
    added, removed, changed = compare_firewall_policies(df_before, df_after)
    display_and_save_results(added, removed, changed, output_filename)