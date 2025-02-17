import pandas as pd
pd.options.mode.chained_assignment = None
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
import re
from datetime import datetime, timedelta
import os

COLUMNS = [
    'Rule Name', 'Source', 'User', 'Destination', 'Service', 'Application', 'Description',
    'REQUEST_ID', 'REQUEST_START_DATE', 'REQUEST_END_DATE', 'TITLE', 'REQUESTER_ID',
    'REQUESTER_EMAIL', 'REQUESTER_NAME', 'REQUESTER_DEPT', 'WRITE_PERSON_ID', 'WRITE_PERSON_EMAIL',
    'WRITE_PERSON_NAME', 'WRITE_PERSON_DEPT', 'APPROVAL_PERSON_ID', 'APPROVAL_PERSON_EMAIL',
    'APPROVAL_PERSON_NAME', 'APPROVAL_PERSON_DEPT_NAME'
]

COLUMNS_NO_HISTORY = [
    'Rule Name', 'Source', 'User', 'Destination', 'Service', 'Application', 'Description'
]

DATE_COLUMNS = ['REQUEST_START_DATE', 'REQUEST_END_DATE']

TRANSLATED_COLUMNS = {
    'REQUEST_ID': 'GSAMS 신청번호',
    'REQUEST_START_DATE': '시작일',
    'REQUEST_END_DATE': '종료일',
    'TITLE': '제목',
    'REQUESTER_ID': '신청자 ID',
    'REQUESTER_EMAIL': '신청자 이메일',
    'REQUESTER_NAME': '신청자명',
    'REQUEST_DEPT': '신청자 부서',
    'WRITE_PERSON_ID': '기안자 ID',
    'WRITE_PERSON_EMAIL': '기안자 이메일',
    'WRITE_PERSON_NAME': '기안자명',
    'WRITE_PERSON_DEPT': '기안자 부서',
    'APPROVAL_PERSON_ID': '결재자 ID',
    'APPROVAL_PERSON_EMAIL': '결재자 이메일',
    'APPROVAL_PERSON_NAME': '결재자명',
    'APPROVAL_PERSON_DEPT': '결재자 부서',
}

except_list = [
    '1',
    '2',
    '3',
]

def update_version(filename: str, final_version: bool = False) -> str:
    base_name, ext = filename.rsplit('.', 1)

    match = re.search(r'_v(\d+)$', base_name)
    final_match = re.search(r'_vf$', base_name)

    if final_match:
        return filename
    
    if final_version:
        if match:
            new_base_name = re.sub(r'_v\d+$', '_vf', base_name)
        else:
            new_base_name = f"{base_name}_vf"
    else:
        if match:
            version = int(match.group(1))
            new_version = version + 1
            new_base_name = re.sub(r'_v\d+$', f'_v{new_version}', base_name)
        else:
            new_base_name = f"{base_name}_v1"
    
    new_filename = f"{new_base_name}.{ext}"
    return new_filename

def select_xlsx_files(extension='.xlsx'):
    file_list = [file for file in os.listdir() if file.endswith(extension)]
    if not file_list:
        print("no excel file")
        return None
    
    for i ,file in enumerate(file_list, start=1):
        print(f"{i}. {file}")
    
    while True:
        choice = input("input file number (exit: 0) ")
        if choice.isdigit():
            choice = int(choice)
            if choice == 0:
                print('exit the program')
                return None
            elif 1 <= choice <= len(file_list):
                return file_list[choice -1]
        print('Invalid number. try again.')
    
def save_to_excel(df, type, file_name):
    wb = load_workbook(file_name)
    sheet = wb[type]

    sheet.insert_rows(1)
    sheet['A1'] = '="대상 정책 수: "&COUNTA(B:B)-1'

    sheet['A1'].font = Font(bold=True)

    for col in range(1, 8):
        cell = sheet.cell(row=2, column=col)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    if type != '이력없음_미사용정책':
        for col in range(8, 24):
            cell = sheet.cell(row=2, column=col)
            cell.fill = PatternFill(start_color='ccffff', end_color='ccffff', fill_type='solid')
    
    wb.save(file_name)

def remove_extension(filename):
    return os.path.splitext(filename)[0]

# 1. parsing request id from description
def parse_request_type():
    def convert_to_date(date_str):
        try:
            date_obj = datetime.strptime(date_str, '%Y%m%d')
            return date_obj.strftime('%Y-%m-%d')
        except ValueError:
            return date_str
    
    def parse_request_info(rulename, description):
        data_dict = {
            "Request Type": "Unknown",
            "Request ID": None,
            "Ruleset ID": None,
            "MIS ID": None,
            "Request User": None,
            "Start Date": convert_to_date('19000101'),
            "End Date": convert_to_date('19000101'),
        }

        if pd.isnull(description):
            return data_dict
        
        # initial pattern
        pattern_gsams = re.compile("MASKED")
        pattern_1_rulename = re.compile("MASKED")
        pattern_1_user = r'MASKED'
        rulename_1_rulename = r'MASKED'
        rulename_1_date = r'MASKED'

        # matched
        match_3 = pattern_3.match(description)
        name_match = pattern_1_rulename.match(str(rulename))
        user_match = re.search(pattern_1_user, description)
        desc_match = re.search(rulename_1_rulename, description)
        date_match = re.search(rulename_1_date, description)

        if match_3:
            data_dict = {
                "Request Type": None,
                "Request ID": match_3.group(5),
                "Ruleset ID": match_3.group(1),
                "MIS ID": match_3.group(6) if match_3.group(6) else None,
                "Request User": match_3.group(4),
                "Start Date": convert_to_date(match_3.group(2)),
                "End Date": convert_to_date(match_3.group(3)),
            }
            
            type_code = data_dict["Request ID"][:1]
            if type_code == "P":
                data_dict["Request Type"] = "GROUP"
            elif type_code == "F":
                data_dict["Request Type"] = "NORMAL"
            elif type_code == "S":
                data_dict["Request Type"] = "SERVER"
            elif type_code == "M":
                data_dict["Request Type"] = "PAM"
            else:
                data_dict["Request Type"] = "Unknown"
        
        if name_match:
            data_dict['Request Type'] = "OLD"
            data_dict['Request ID'] = name_match.group(1)
            if user_match:
                data_dict['Request User'] = user_match.group(1).replace("*ACL*", "")
            if date_match:
                data_dict['Start Date'] = convert_to_date(date_match.group().split("~")[0])
                data_dict['End Date'] = convert_to_date(date_match.group().split("~")[1])
        
        if desc_match:
            date = description.split(';')[0]
            start_date = date.split('~')[0].replace('[', '').replace('-', '')
            end_date = date.split('~')[1].replace(']', '').replace('-', '')

            date_dict = {
                "Request Type": "OLD",
                "Request ID": desc_match.group(1).split('-')[1],
                "Ruleset ID": None,
                "MIS ID": None,
                "Request User": user_match.group(1).replace("*ACL*", "") if user_match else None,
                "Start Date": convert_to_date(start_date),
                "End Date": convert_to_date(end_date),
            }
        
        return data_dict
    
    file_name = select_xlsx_files()
    df = pd.read_excel(file_name)

    for index, row in df.iterrows():
        result = parse_request_info(row['Rule Name'], row['Description'])
        for key, value in result.items():
            df.at[index, key] = value
    
    df.to_excel(update_version(file_name), index=False)

# 2. extract request id
def extract_request_id():
    file_name = select_xlsx_files()
    df = pd.read_excel(file_name)

    # 'Unknown' 값을 제외하고 고유한 Request Type 값을 추출
    unique_types = df[df['Request Type'] != 'Unknown']['Request Type'].unique()

    # 고유한 Request Type 값을 최대 5개 선택
    selected_types = unique_types[:5]

    # 선택된 Request Type에 해당하는 데이터 추출
    selected_data = df[df['Request Type'].isin(selected_types)]

    # 각 Request Type별로 Request ID 값만 추출하여 중복 제거 후 Excel의 각 시트로 저장
    with pd.ExcelWriter(f"request_id_{file_name}") as writer:
        for request_type, group in selected_data.groupby('Request Type'):
            group[['Request ID']].drop_duplicates().to_excel(writer, sheet_name=request_type, index=False)

# 3. add request info
def add_request_info():
    def read_and_process_excel(file):
        """ Excel 파일 읽기 및 초기 처리 """
        df = pd.read_excel(file)
        df.replace({'nan': None}, inplace=True)
        return df.astype(str)
    
    def match_and_update_df(rule_df, info_df):
        """ 조건에 따라 DataFrame의 값을 매칭 및 업데이트 """
        total = len(rule_df)
        for idx, row in rule_df.iterrows():
            print(f"\rProgress: {idx +1}/{total}", end='', flush=True)
            if row['Request Type'] == 'GROUP':
                matched_row = info_df[
                    ((info_df['REQUEST_ID'] == row['Request ID']) & (info_df['MIS_ID'] == row['MIS ID'])) |
                    ((info_df['REQUEST_ID'] == row['Request ID']) & (info_df['REQUEST_END_DATE'] == row['End Date']) & (info_df['WRITE_PERSON_ID'] == row['Request User'])) |
                    ((info_df['REQUEST_ID'] == row['Request ID']) & (info_df['REQUEST_END_DATE'] == row['End Date']) & (info_df['REQUESTER_ID'] == row['Request User']))
                ]
            else:
                matched_row = info_df[info_df['REQUEST_ID'] == row['Request ID']]
            
            if not matched_row.empty:
                for col in matched_row.columns:
                    if col in ['REQUEST_START_DATE', 'REQUEST_END_DATE', 'Start Date', 'End Date']:
                        rule_df.at[idx, col] = pd.to_datetime(matched_row[col].values[0], errors='coerce')
                    else:
                        rule_df.at[idx, col] = matched_row[col].values[0]
            elif row['Request Type'] != 'nan' and row['Request Type'] != 'Unknown':
                rule_df.at[idx, 'REQUEST_ID'] = row['Request ID']
                rule_df.at[idx, 'REQUEST_START_DATE'] = row['Start Date']
                rule_df.at[idx, 'REQUEST_END_DATE'] = row['End Date']
                rule_df.at[idx, 'REQUESTER_ID'] = row['Request User']
                rule_df.at[idx, 'REQUESTER_EMAIL'] = row['Request User'] + '@gmail.com'

    print('select policy file: ')
    rule_file = select_xlsx_files()

    if not rule_file:
        return False
    
    print("select info file: ")
    info_file = select_xlsx_files()
    if not info_file:
        return False
    
    rule_df = read_and_process_excel(rule_file)
    info_df = read_and_process_excel(info_file)
    info_df = info.sort_values(by='REQUEST_END_DATE', ascending=False)
    auto_extension_id = find_auto_extension_id()
    match_and_update_df(rule_df, info_df)
    rule_df.replace({'nan': None}, inplace=True)

    rule_df.loc[rule_df['REQUEST_ID'].isin(auto_extension_id), 'REQUEST_STATUS'] = '99'

    rule_df.to_excel(update_version(rule_file), index=False)

# 4. exception pa
def paloalto_exception():
    print("seelct policy file: ")
    rule_file = select_xlsx_files()
    df = pd.read_excel(rule_file)

    current_date = datetime.now()
    three_months_ago = current_date - timedelta(days=90)

    df["예외"] = ''

    # 1. except list와 request id 일치 시 예외 신청정책으로 표시
    df['REQUEST_ID'] = df['REQUEST_ID'].fillna('')
    for id in except_list:
        df.loc[df['REQUEST_ID'].str.startswith(id, na=False), '예외'] = '예외신청정책'
    
    # 2.
    df.loc[df['REQUEST_STATUS'] == 99, '예외'] = '자동연장정책'

    # 3.
    df['날짜'] = df['Rule Name'].str.extract(r'(\d{8})', expand=False)
    df['날짜'] = pd.to_datetime(df['날짜'], format='%Y%m%d', errors='coerce')
    df.loc[(df['날짜'] >= three_months_ago) & (df['날짜'] <= current_date), '예외'] = '신규정책'

    # 4.
    deny_std_rule_index = df[df['Rule Nam'] == 'deny_rule'].index[0]
    df.log[df.index < deny_std_rule_index, '예외'] = '인프라정책'

    # 5.
    df.loc[df['Rule Name'].str.startswith(('sample_', 'test_')), '예외'] = 'test_group_정책'

    # 6.
    df.loc[df['Enable'] == 'N', '예외'] = '비활성화정책'

    # 7.
    df.loc[(df['Rule Name'].str.endswith('_Rule')) & (df['Enable'] == 'N'), '예외'] = '기준정책'

    # 8.
    df.log[df['Action'] == 'deny', '예외'] = '차단정책'

    df['예외'].fillna('', inplace=True)

    cols = list(df.columns)
    cols = ['예외'] + [col for col in cols if col != '예외']
    df = df[cols]

    def check_date(row):
        try:
            end_date = pd.to_datetime(row['REQUEST_END_DATE'])
            return '미만료' if end_date > current_date else '만료'
        except:
            return '만료'
    

    df['만료여부'] = df.apply(check_date, axis=1)

    df.drop(columns=['날짜'], inplace=True)

    df.rename(columns={'Request Type': '신청이력'}, inplace=True)

    df.drop(columns=['Request ID', 'Ruleset ID', 'MIS ID', 'Request User', 'Start Date', 'End Date'], inplace=True)

    cols = list(df.columns)
    cols.insert(cols.index('예외') + 1, cols.pop(cols.index('만료여부')))
    df = df[cols]

    cols.insert(cols.index('예외') + 1, cols.pop(cols.index('신청이력')))
    df = df[cols]

    cols.insert(cols.index('만료여부') + 1, '미사용여부')
    df['미사용여부'] = ''

    df.to_excel(update_version(rule_file, True), index=False, engine='openpyxl')

# 5. exception secui
def secui_exception():
    print("seelct policy file: ")
    rule_file = select_xlsx_files()
    df = pd.read_excel(rule_file)

    current_date = datetime.now()
    three_months_ago = current_date - timedelta(days=90)

    df["예외"] = ''

    # 1. except list와 request id 일치 시 예외 신청정책으로 표시
    df['Request ID'].fillna('-', inplace=True)

    for id in except_list:
        df.loc[df['Request ID'].str.startswith(id), '예외'] = '예외신청정책'
    
    # 2.
    df.loc[df['REQUEST_STATUS'] == 99, '예외'] = '자동연장정책'

    # 4.
    deny_std_rule_index = df[df['Description'].str.contains('deny_rule') == True].index[0]
    df.log[df.index < deny_std_rule_index, '예외'] = '인프라정책'

    # 5.
    df.log[df['Description'].str.contains(('sample_', 'test_')) ==True, '예외'] = 'test_group_정책'

    # 6.
    df.loc[df['Enable'] == 'N', '예외'] = '비활성화정책'

    # 7.
    df.loc[(df['Description'].str.contains('기준룰')) & (df['Enable'] == 'N'), '예외'] = '기준정책'

    # 8.
    df.log[df['Action'] == 'deny', '예외'] = '차단정책'

    df['예외'].fillna('', inplace=True)

    cols = list(df.columns)
    cols = ['예외'] + [col for col in cols if col != '예외']
    df = df[cols]

    def check_date(row):
        try:
            end_date = pd.to_datetime(row['REQUEST_END_DATE'])
            return '미만료' if end_date > current_date else '만료'
        except:
            return '만료'
    

    df['만료여부'] = df.apply(check_date, axis=1)

    df.rename(columns={'Request Type': '신청이력'}, inplace=True)

    df.drop(columns=['Request ID', 'Ruleset ID', 'MIS ID', 'Request User', 'Start Date', 'End Date'], inplace=True)

    cols = list(df.columns)
    cols.insert(cols.index('예외') + 1, cols.pop(cols.index('만료여부')))
    df = df[cols]

    cols.insert(cols.index('예외') + 1, cols.pop(cols.index('신청이력')))
    df = df[cols]

    cols.insert(cols.index('만료여부') + 1, '미사용여부')
    df = df.reindex(columns=cols)
    df['미사용여부'] = ''

    df.to_excel(update_version(rule_file, True), index=False, engine='openpyxl')

def find_auto_extension_id():
    print('가공된 신청정보 파일을 선택')
    selected_file = select_xlsx_files()
    df = pd.read_excel(selected_file)
    filtered_df = df[df['REQEUST_STATUS'].isin([98, 99])]['REQUEST_ID'].drop_duplicates()

    return filtered_df

def organize_redundant_file():
    expected_columns = ['No', 'Type', 'Seq', 'Rule Name', 'Enable', 'Action', 'Source', 'User', 'Destination', 'Service', 'Application', 'Security Profile', 'Category', 'Description', 'Request Type', 'Request ID', 'Ruleset ID', 'MIS ID', 'Request User', 'Start Date', 'End Date']
    expected_columns_2 = ['No', 'Type', 'Vsys', 'Seq', 'Rule Name', 'Enable', 'Action', 'Source', 'User', 'Destination', 'Service', 'Application', 'Security Profile', 'Category', 'Description', 'Request Type', 'Request ID', 'Ruleset ID', 'MIS ID', 'Request User', 'Start Date', 'End Date']

    try:
        print('중복정책 파일을 선택')
        selected_file = select_xlsx_files()
        df = pd.read_excel(selected_file)

        auto_extension_id = find_auto_extension_id()

        current_columns = df.columns.tolist()

        if current_columns != expected_columns or current_columns != expected_columns_2:
            print('컬럼명 일치')
        else:
            print('컬럼명 불일치')
    
    except Exception as e:
        print('엑셀 파일을 열 수 없습니다.')
        print(e)
        exit()
    
    df['자동연장'] = df['Request ID'].isin(auto_extension_id)

    df['늦은종료일'] = df.groupby('No')['End Date'].transform(lambda x: (x == x.max()) & (~x.duplicated(keep='first')))

    df['신청자검증'] = df.groupby('No')['Request User'].transform(lambda x: x.nunique() == 1)

    target_rule_true = df[(df['Type'] == 'Upper') & (df['늦은종료일'] == True)]['No'].unique()

    df['날짜검증'] = False

    df.loc[df['No'].isin(target_rule_true), '날짜검증'] = True

    df['작업구분'] = '유지'
    df.loc[df['늦은종료일'] == False, '작업구분'] = '삭제'

    df['공지여부'] = False
    df.loc[df['신청자검증'] == False, '공지여부'] = True

    df['미사용예외'] = False
    df.loc[(df['날짜검증'] == False) & (df['늦은종료일'] == True), '미사용예외'] = True

    extensioned_df = df.groupby('No').filter(lambda x: x['자동연장'].any())
    extensioned_group = extensioned_df[extensioned_df['Request Type'] == 'GROUP']
    exception_target = extensioned_group.groupby('No').filter(lambda x: len(x['Request ID'].unique()) >=2 )
    exception_id = exception_target[(exception_target['자동연장'] == True) & (exception_target['작업구분'] == '삭제')]['No']

    df = df[~df['No'].isin(exception_id)]

    filtered_no = df.groupby('No').filter(
        lambda x: (x['Request Type'] != 'GROUP').any() and
                (x['작업구분'] == '삭제').any() and
                (x['자동연장'] == True).any()
    )['No'].unique()

    df = df[~df['No'].isin(filtered_no)]

    filtered_no_2 = df.groupby('No').filter(
        lambda x: (x['작업구분'] != '유지').all()
    )['No'].unique()

    df = df[~df['No'].isin(filtered_no_2)]

    target_types = ["PAM", "SERVER", "Unknown"]
    target_nos = df[df['Request Type'].isin(target_types)]['No'].drop_duplicates()
    
    df = df[~df['No'].isin(target_nos)]

    notice_df = df[df['공지여부'] == True]
    delete_df = df[df['공지여부'] == False]

    column_to_move = notice_df.pop('작업구분')
    notice_df.insert(0, '작업구분', column_to_move)
    column_to_move = delete_df.pop('작업구분')
    delete_df.insert(0, '작업구분', column_to_move)
    
    notice_df.drop(['Request Type', 'Ruleset ID', 'MIS ID', 'Start Date', 'End Date', '늦은종료일', '신청자검증', '날짜검증', '공지여부', '미사용예외', '자동연장'], axis=1, inplace=True)
    delete_df.drop(['Request Type', 'Ruleset ID', 'MIS ID', 'Start Date', 'End Date', '늦은종료일', '신청자검증', '날짜검증', '공지여부', '미사용예외', '자동연장'], axis=1, inplace=True)
    
    filename = remove_extension(selected_file)
    output_excel_path = f'{filename}_정리.xlsx'
    notice_excel_path = f'{filename}_공지.xlsx'
    delete_excel_path = f'{filename}_삭제.xlsx'

    df.to_excel(output_excel_path, index=False, engine='openpyxl')
    notice_df.to_excel(notice_excel_path, index=False, engine='openpyxl')
    delete_df.to_excel(delete_excel_path, index=False, engine='openpyxl')

def add_mis_id():
    print("select policy file")
    file = select_xlsx_files()
    print("select mis id file")
    mis_df = pd.read_csv(select_xlsx_files(".csv"))
    rule_df = pd.read_excel(file)

    mis_df_unique = mis_df.drop_duplicates(subset=['ruleset_id'], keep='first')

    mis_id_map = mis_df_unique.set_index('ruleset_id')['mis_id']

    rule_df['MIS ID'] = rule_df.apply(lambda row: mis_id_map.get(row['Ruleset ID'], row['MIS ID']) if pd.isna(row['MIS ID']) or row['MIS ID'] == '' else row['MIS ID'], axis=1)

    rule_df.to_excel(update_version(file, False), index=False, engine='openpyxl')

def notice_file_organization():
    def expired_used(df, selected_file):
        filtered_df = df[
            ((df['예외'].isna()) | (df['예외'] == '신규정책')) &
            (df['중복여부'].isna()) &
            (df['신청이력'] != 'Unknown') &
            (df['만료여부'] == '만료') &
            (df['미사용여부'] == '사용')
        ]

        selected_df = filtered_df[COLUMNS]
        selected_df = selected_df.astype(str)

        for date_column in DATE_COLUMNS:
            selected_df[date_column] = pd.to_datetime(selected_df[date_column]).dt.strftime('%Y-%m-%d')
        
        selected_df.rename(TRANSLATED_COLUMNS, inplace=True)
        selected_df.fillna('', inplace=True)
        selected_df.replace('nan', '', inplace=True)
        
        type = '만료_사용정책'
        filename = str(remove_extension(selected_file)) + '_기간만료(공지용).xlsx'
        selected_df.to_excel(filename, index=False, na_rep='', sheet_name=type)
        save_to_excel(selected_df, type, filename)

    def expired_unused(df, selected_file):
        filtered_df = df[
            ((df['예외'].isna()) | (df['예외'] == '신규정책')) &
            (df['중복여부'].isna()) &
            (df['신청이력'] != 'Unknown') &
            (df['만료여부'] == '만료') &
            (df['미사용여부'] == '미사용')
        ]

        selected_df = filtered_df[COLUMNS]
        selected_df = selected_df.astype(str)

        for date_column in DATE_COLUMNS:
            selected_df[date_column] = pd.to_datetime(selected_df[date_column]).dt.strftime('%Y-%m-%d')
        
        selected_df.rename(TRANSLATED_COLUMNS, inplace=True)
        selected_df.fillna('', inplace=True)
        selected_df.replace('nan', '', inplace=True)

        type = '만료_미사용정책'
        filename = str(remove_extension(selected_file)) + '_만료_미사용정책(공지용).xlsx'
        selected_df.to_excel(filename, index=False, na_rep='', sheet_name=type)
        save_to_excel(selected_df, type, filename)

    def longterm_unused_rules(df, selected_file):
        filtered_df = df[
            (df['예외'].isna()) &
            (df['중복여부'].isna()) &
            (df['신청이력'] != 'Unknown') &
            (df['만료여부'] == '미만료') &
            (df['미사용여부'] == '미사용')
        ]

        selected_df = filtered_df[COLUMNS]
        selected_df = selected_df.astype(str)

        for date_column in DATE_COLUMNS:
            selected_df[date_column] = pd.to_datetime(selected_df[date_column]).dt.strftime('%Y-%m-%d')
        
        selected_df.rename(TRANSLATED_COLUMNS, inplace=True)
        selected_df.fillna('', inplace=True)
        selected_df.replace('nan', '', inplace=True)

        type = '미만료_미사용정책'
        filename = str(remove_extension(selected_file)) + '_장기미사용정책(공지용).xlsx'
        selected_df.to_excel(filename, index=False, na_rep='', sheet_name=type)
        save_to_excel(selected_df, type, filename)

    def no_history_unused(df, selected_file):
        filtered_df = df[
            (df['예외'].isna()) &
            (df['중복여부'].isna()) &
            (df['신청이력'] != 'Unknown') &
            (df['만료여부'] == '만료') &
            (df['미사용여부'] == '미사용')
        ]

        selected_df = filtered_df[COLUMNS_NO_HISTORY]
        selected_df = selected_df.astype(str)

        selected_df.fillna('', inplace=True)
        selected_df.replace('nan', '', inplace=True)

        type = '이력없음_미사용정책'
        filename = str(remove_extension(selected_file)) + '_이력없는_미사용정책.xlsx'
        selected_df.to_excel(filename, index=False, na_rep='', sheet_name=type)
        save_to_excel(selected_df, type, filename)

    print("분류할 정책파일을 선택하세요.")
    selected_file = select_xlsx_files()
    logging.info("정책 분류 시작")
    try:
        df = pd.read_excel(selected_file)
        try:
            expired_used(df, selected_file)
            logging.info("기간만료 분류 완료")
        except:
            logging.error("기간만료 분류 실패")
        try:
            expired_unused(df, selected_file)
            logging.info("만료/미사용 분류 완료")
        except:
            logging.error("만료/미사용 분류 실패")
        try:
            longterm_unused_rules(df, selected_file)
            logging.info("장기미사용 분류 완료")
        except:
            logging.error("장기미사용 분류 실패")
        try:
            no_history_unused(df, selected_file)
            logging.info("이력없는 미사용 분류 완료")
        except:
            logging.error("이력없는 미사용 분류 실패")
        logging.info("정책 분류 완료")
    except:
        logging.error("정책 분류 실패")

# def app_info_processing():
#     def parse_group_df(df):
#         expected_columns = [
#             '신청번호',
#             '시작일',
#             '종료일',
#             '제목',
#             'REQUEST_STATUS',
#             '진행상태',
#             '신청자 ID',
#             '신청자 이메일',
#             '신청자명',
#             '신청자 부서',
#             '기안자 ID',
#             '기안자 이메일',
#             '기안자명',
#             '기안자 부서',
#             '결재자ID',
#             '결재자 이메일',
#             '결재자명',
#             '결재자 부서',
#             'REQUEST_DATE',
#             'MIS_ID'
#         ]

#         actual_columns = df.columns.tolist()

#         if not actual_columns == expected_columns:
#             exit("컬럼명 불일치")
        
#         df = df.drop(columns=['진행상태'])

#         cols = df.columns.tolist()
#         columns_to_move = 'REQUEST_STATUS'
#         insert_before = 'MIS_ID'

#         cols.remove(columns_to_move)
#         insert_index = cols.index(insert_before)
        
#         cols.insert(insert_index, columns_to_move)
#         df = df[cols]

#         columns_to_rename = {

#         }

def select_task():
    print("시작할 작업 번호를 입력해주세요.")
    print("1. Description에서 신청번호 파싱하기")
    print("2. 정책파일에서 신청번호 추출하기")
    print("3. 정책파일에서 신청정보 추가하기")
    print("4. 팔로알토 정책에서 예외처리하기")
    print("5. 시큐아이 정책에서 예외처리하기")
    print("6. 중복정책 공지/삭제 분류하기")
    print("7. 정리대상 별 공지파일 분류하기")
    print("8. 정책파일에서 MIS ID 추가하기")
    print("0. 종료")

    while True:
        try:
            choice = int(input("작업 번호 (1-9, 종료: 0)"))
            if 1 <= choice <= 9:
                return choice
            elif choice == 0:
                exit('exit')
            else:
                print('invalid number')
        except ValueError:
            print("Invalid input")

def deletion_process_main():
    start_task = select_task()

    try:
        logging.info(f"Starting deletion process")
        if start_task == 1:
            parse_request_type()
        elif start_task <= 2:
            extract_request_id()
        elif start_task <= 3:
            add_request_info()
        elif start_task <= 4:
            paloalto_exception()
        elif start_task <= 5:
            secui_exception()
        elif start_task <= 6:
            organize_redundant_file()
        elif start_task <= 7:
            notice_file_organization()
        elif start_task <= 8:
            add_mis_id()
        else:
            logging.info("Exiting deletion process")
            exit()
        logging.info("Completed deletion process")
    
    except Exception as e:
        logging.exception(f"Exception in deletion process : {e}")