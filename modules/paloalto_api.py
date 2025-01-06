import requests
requests.packages.urllib3.util.ssl_.DEFAULT_CIPHERS += ':DES-CBC3-SHA'
requests.packages.urllib3.disable_warnings()
import xml.etree.ElementTree as ET
import pandas as pd
import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
import time
import os
import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def apply_excel_style(file_name):
    """
        주어진 엑셀파일의 헤더에 연한 회색 배경을 적용하고,
        헤더의 너비를 자동으로 조절하되 최대 너비를 40으로 제한하는 함수.

        :param file_name: 처리할 엑셀파일의 이름
    """
    try:
        # load excel file
        workbook = load_workbook(file_name)
        worksheet = workbook.active

        # set range of header (first row is header)
        header_range = worksheet[1]

        fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        for cell in header_range:
            cell.fill = fill
            column_letter = cell.column_letter
            max_length = 0
            for cell in worksheet[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = min(40, adjusted_width)
        
        workbook.save(file_name)
    
    except Exception as e:
        print(f"An error occurred: {e}")

class PaloAltoAPI:
    def __init__(self, hostname, username, password):
        self.hostname = hostname
        self.base_url = f'https://{hostname}/api/'
        self.api_key = self.get_api_key(username, password)

    def save_dfs_to_excel(self, dfs, sheet_names, file_name):
        try:
            if not isinstance(dfs, list):
                dfs = [dfs]
            
            if not isinstance(sheet_names, list):
                sheet_names = [sheet_names]
            
            with pd.ExcelWriter(file_name) as writer:
                for df, sheet_name in zip(dfs, sheet_names):
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_excel_style(file_name)
            return True
        except:
            return False
    
    def save_df_to_excel(self, df, df_name):
        current_date = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
        save_file = f"{current_date}_{self.hostname}_{df_name}.xlsx"
        if os.path.isfile(save_file):
            raise FileExistsError(f"File '{save_file}' already exists")
        
        with pd.ExcelWriter(save_file, engine='openpyxl') as writer:
            df.to_excel(wirter, sheet_name=df_name, index=False)
    
    def get_member(self, entry):
        try:
            result = [ member.text for member in entry ]
        except:
            result = []
        
        return result
    
    def list_to_string(self, list_data):
        return ','.join(str(s) for s in list_data)
    
    def get_api_data(self, parameter: dict, time_out: int = 10000):
        try:
            response = requests.get(self.base_url, params=parameter, verify=False, timeout=time_out)
            return response
        
        except requests.exceptions.RequestException as e:
            raise ValueError(f"API 요청 중 요류 발생: {e}")
    
    def get_api_key(self, username:str, password: str):
        keygen_parameter = (
            ('type', 'keygen'),
            ('user', username),
            ('password', password)
        )

        try:
            response = self.get_api_data(keygen_parameter)
            key_value = ET.fromstring(response.text).find('./result/key')
            api_key = key_value.text
            
            return api_key
        
        except requests.exceptions.RequestException as e:
            raise ValueError(f"API 요청 중 오류 발생: {e}")
    
    def get_vsys_list(self):
        parameter = (
            ('key', self.api_key),
            ('type', 'config'),
            ('action', 'get'),
            ('xpath', '/config/devices/entry/vsys/entry'),
        )

        response = self.get_api_data(parameter)
        root = ET.fromstring(response.text).findall('./result/entry')
        vsys_list = []
        for vsys in root:
            vsys_list.append(vsys.attrib.get('name'))

        return vsys_list
    
    def get_config(self, config_type: str = 'running'):
        action = 'show' if config_type == 'running' else 'get'
        parameter = (
            ('key', self.api_key),
            ('type', 'config'),
            ('action', action),
            ('xpath', '/config')
        )

        response = self.get_api.data(parameter)

        return response.text
    
    def save_config(self, config_type: str = 'running'):
        current_date = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
        config = self.get_config(config_type)
        with open(f'{current_date}_{self.hostname}_{config_type}_config.xml', mode='w', encoding='utf8') as file:
            file.write(config)
        
        return True
    
    def get_system_info(self):
        parameter = (
            ('type', 'op'),
            ('cmd', '<show><system><info/></system></show>'),
            ('key', self.api_key)
        )
        response = self.get_api_data(parameter)

        tree = ET.fromstring(response.text)
        info = {
            "hostname": tree.findtext("./result/system/hostname"),
            "ip_address": tree.findtext("./result/system/ip-address"),
            "mac_address": tree.findtext("./result/system/mac-address"),
            "uptime": tree.findtext("./result/system/uptime").split(" ")[0],
            "model": tree.findtext("./result/system/model"),
            "serial_number": tree.findtext("./result/system/serial"),
            "sw_version": tree.findtext("./result/system/sw-version"),
            "app_version": tree.findtext("./result/system/app-version"),
        }

        return pd.DataFrame(info, index=[0])
    
    def get_system_state(self):
        parameter = (
            ('type', 'op'),
            ('cmd', '<show><system><state><filter>cfg.general.max*</filter></state></system></show>'),
            ('key', self.api_key)
        )
        response = self.get_api_data(parameter)
        tree = ET.fromstring(response.text)

        result = tree.findtext("./result")
        for data in result.split('\n'):
            if data.startswith('cfg.general.max-address:'):
                max_address = data.split(': ')[1]
            elif data.startswith('cfg.general.max-address-group:'):
                max_address_group = data.split(': ')[1]
            elif data.startswith('cfg.general.max-service:'):
                max_service = data.split(': ')[1]
            elif data.startswith('cfg.general.max-service-group:'):
                max_service_group = data.split(': ')[1]
            elif data.startswith('cfg.general.max-policy-rule:'):
                max_policy_rule = data.split(': ')[1]
        
        state = {
            "hostname": self.hostname,
            "max_policy_rule": max_policy_rule,
            "max_address": max_address,
            "max_address_group": max_address_group,
            "max_service": max_service,
            "max_service_group": max_service_group
        }

        return pd.DataFrame(state, index=[0])
    
    def export_security_rules(self, config_type: str = 'running'):
        config = self.get_config(config_type)
        tree = ET.fromstring(config)
        vsys_list = tree.findall('./result/config/devices/entry/vsys/entry')
        security_rules = []

        for vsys in vsys_list:
            vsys_name = vsys.attrib.get('name')
            rulebase = vsys.findall('./rulebase/security/rules/entry')
            for idx, rule in enumerate(rulebase):
                rule_name = str(rule.attrib.get('name'))
                disabled = self.list_to_string(self.get_member(rule.findall('./disabled')))
                disabled = "N" if disabled == "yes" else "Y"
                action = self.list_to_string(self.get_member(rule.findall('./action')))
                source_list = self.list_to_string(self.get_member(rule.findall('./source/member')))
                source_user_list = self.list_to_string(self.get_member(rule.findall('./source-user/member')))
                destination_list = self.list_to_string(self.get_member(rule.findall('./destination/member')))
                service_list = self.list_to_string(self.get_member(rule.findall('./service/member')))
                application_list = self.list_to_string(self.get_member(rule.findall('./application/member')))
                url_filtering = self.list_to_string(self.get_member(rule.findall('./profile-setting/profiles/url-filtering/member')))
                category = self.list_to_string(self.get_member(rule.findall('./category/member')))
                category = "any" if category == "" else category
                description = self.get_member(rule.findall('./description'))
                description = self.list_to_string([ i.replace('\n',' ') if type(i) is str else i for i in description ])

                rule_info = {
                    "Vsys": vsys_name,
                    "Seq": idx+1,
                    "Rule Name": rule_name,
                    "Enable": disabled,
                    "Action": action,
                    "Source": source_list,
                    "User": source_user_list,
                    "Destination": destination_list,
                    "Service": service_list,
                    "Application": application_list,
                    "Security Profile": url_filtering,
                    "Category": category,
                    "Description": description,
                }
                security_rules.append(rule_info)
        
        return pd.DataFrame(security_rules)
    
    def export_network_objects(self, config_type: str = 'running'):
        config = self.get_config(config_type)
        tree = ET.fromstring(config)
        address_list = tree.findall('./result/config/devices/entry/vsys/entry/address/entry')
        address_objects = []
        for address in address_list:
            ip = address.attrib.get('name')
            ip_name = ip
            ip_member = []
            address_type = address.find('*').tag
            ip_entry = address.findall(f'./{address_type}')
            for j in ip_entry:
                ip_member.append(j.text)
            
            objects_info = {
                "Name": ip_name,
                "Type": address_type,
                "Value": self.list_to_string(ip_member)
            }
            address_objects.append(objects_info)
        
        return pd.DataFrame(address_objects)
    
    def export_network_group_objects(self, config_type: str = 'running'):
        config = self.get_config(config_type)
        tree = ET.fromstring(config)
        address_group_list = tree.findall('./result/config/devices/entry/vsys/entry/address-group/entry')
        address_group_objects = []
        for address_group in address_group_list:
            group = address_group.attrib.get('name')
            group_name = group
            group_member = []
            
            group_entry = address_group.findall('./static/member')
            for j in group_entry:
                group_member.append(j.text)

            objects_info = {
                "Group Name": group_name,
                "Entry": self.list_to_string(group_member)
            }
            address_group_objects.append(objects_info)
        
        return pd.DataFrame(address_group_objects)
    
    def export_service_objects(self, config_type: str = 'running'):
        config = self.get_config(config_type)
        tree = ET.fromstring(config)
        services = tree.findall('./result/config/devices/entry/vsys/entry/service/entry')
        service_objects = []
        for service in services:
            service_info = service.attrib.get('name')
            service_name = service_info

            for protocol in service.find('protocol'):
                protocol_name = protocol.tag
                port = protocol.find('port').text if protocol.find('port') is not None else None

                service_info = {
                    "Name": service_name,
                    "Protocol": protocol_name,
                    "Port": port,
                }
                service_objects.append(service_info)
        
        return pd.DataFrame(service_objects)
    
    def export_service_group_objects(self, config_type: str = 'running'):
        config = self.get_config(config_type)
        tree = ET.fromstring(config)
        service_groups = tree.findall('./result/config/devices/entry/vsys/entry/service-group/entry')
        service_group_objects = []
        for service_group in service_groups:
            group_name = service_group.attrib.get('name')
            group_member = []

            group_entry = service_group.findall('./members/member')
            for j in group_entry:
                group_member.append(j.text)
            
            service_ingo = {
                "Group Name": group_name,
                "Entry": self.list_to_string(group_member),
            }

        return pd.DataFrame(service_group_objects)
    
    def export_hit_count(self, vsys_name: str = 'vsys1'):
        parameter = (
            ('type', 'op'),
            ('cmd', f"<show><rule-hit-count><vsys><vsys-name><entry name='{vsys_name}'><rule-base><entry name='security'><rules><all/></rules></entry></rule-base></entry></vsys-name></vsys></rule-hit-count></show>"),
            ('key', self.api_key)
        )

        response = self.get_api.data(parameter)
        tree = ET.fromstring(response.text)
        rules = tree.findall('./result/rule-hit-count/vsys/entry/rule-base/entry/rules/entry')

        result = []
        for rule in rules:
            rule_name = str(rule.attrib.get('name'))
            rule_info = self.get_member(rule)
            hit_count = rule_info[1]
            last_hit_timestamp = int(rule_info[2])
            first_hit_timestamp = int(rule_info[4])

            no_unused_days = 99999
            no_hit_date = datetime.datetime(1900, 1, 1).strftime('%Y-%m-%d')

            if first_hit_timestamp == 0:
                no_unused_days = no_unused_days
            else:
                unused_days = (datetime.datetime.now() - datetime.datetime.fromtimestamp(last_hit_timestamp)).days
            
            if last_hit_timestamp == 0:
                last_hit_date = no_hit_date
            else:
                last_hit_date = datetime.datetime.fromtimestamp(last_hit_timestamp).strftime('%Y-%m-%d')
            
            if first_hit_timestamp == 0:
                first_hit_date = no_hit_date
            else:
                first_hit_date = datetime.datetime.fromtimestamp(first_hit_timestamp).strftime('%Y-%m-%d')
            
            result.append({
                "Vsys": vsys_name,
                "Rule Name": rule_name,
                "Hit Count": hit_count,
                "First Hit Date": first_hit_date,
                "Last Hit Date": last_hit_date,
                "Unused Days": unused_days
            })
        
        return pd.DataFrame(result)