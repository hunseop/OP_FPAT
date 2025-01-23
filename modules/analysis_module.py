import os
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def analyze_redundant_policies(df, vendor):
    """
    Analyzes redundant firewall policies and returns the resulting DataFrame.

    :param df: pandas DataFrame containing firewall policies
    :param vendor: string, vendor type (e.g., 'paloalto')
    :return: pandas DataFrame with analyzed redundant policies
    """

    def ensure_upper_and_lower(df):
        """
        Filters the DataFrame to include only groups where both 'Upper' and 'Lower' exist
        and regenerates the 'No' column.
        """
        valid_no_groups = []
        grouped = df.groupby('No')

        for name, group in grouped:
            if 'Upper' in group['Type'].values and 'Lower' in group['Type'].values:
                valid_no_groups.append(group)

        filtered_df = pd.concat(valid_no_groups).reset_index(drop=True)
        filtered_df['No'] = filtered_df.groupby(['No']).ngroup() + 1
        return filtered_df

    df_filtered = df[(df['Enable'] == 'Y') & (df['Action'] == 'allow')].copy()
    columns_to_check = ['Enable', 'Action', 'Source', 'User', 'Destination', 'Service', 'Application']

    if 'Vsys' in df.columns:
        columns_to_check.append('Vsys')

    if vendor == 'paloalto':
        df_filtered['Service'] = df_filtered['Service'].str.replace('_', '-')
        df_filtered.loc[:, 'Service'] = df_filtered['Service'].str.replace('_', '-')
        columns_to_check.append('Category')

    df_check = df_filtered[columns_to_check]

    def normalize_policy(policy_series):
        normalized_policy = policy_series.apply(lambda x: ','.join(sorted(x.split(','))) if isinstance(x, str) else x)
        return tuple(normalized_policy)

    policy_map = defaultdict(list)
    results_list = []
    current_no = 1

    for i in range(len(df_filtered)):
        current_policy = normalize_policy(df_check.iloc[i])
        if current_policy in policy_map:
            row = df_filtered.iloc[i].to_dict()
            row.update({'No': policy_map[current_policy], 'Type': 'Lower'})
            results_list.append(row)
        else:
            policy_map[current_policy] = current_no
            row = df_filtered.iloc[i].to_dict()
            row.update({'No': current_no, 'Type': 'Upper'})
            results_list.append(row)
            current_no += 1

    results = pd.DataFrame(results_list)

    results = ensure_upper_and_lower(results)

    columns_order = ['No', 'Type'] + [col for col in df.columns]
    results = results[columns_order]
    results = results.sort_values(by=['No', 'Type'], ascending=[True, False])

    return results


def compare_firewall_policies(df_before, df_after):
    """
    Compares two firewall policies and returns summary, added, removed, and changed DataFrames.

    :param df_before: pandas DataFrame containing the "before" state of firewall policies
    :param df_after: pandas DataFrame containing the "after" state of firewall policies
    :return: tuple of pandas DataFrames (summary_df, added_df, removed_df, changed_df)
    """
    df_merged = df_before.merge(
        df_after, on='Rule Name', how='outer', suffixes=('_before', '_after'), indicator=True
    )

    added_df = df_merged[df_merged['_merge'] == 'right_only']
    removed_df = df_merged[df_merged['_merge'] == 'left_only']

    added_cols = ['Rule Name'] + [col for col in added_df.columns if col.endswith('_after')]
    removed_cols = ['Rule Name'] + [col for col in removed_df.columns if col.endswith('_before')]

    added_df = added_df[added_cols].rename(columns=lambda x: x.replace('_after', ''))
    removed_df = removed_df[removed_cols].rename(columns=lambda x: x.replace('_before', ''))

    common_cols = [col for col in df_before.columns if col not in ['Rule Name', 'Seq']]
    change_conditions = [
        df_merged[f'{col}_before'] != df_merged[f'{col}_after'] for col in common_cols
    ]
    changed = df_merged[
        (df_merged['_merge'] == 'both') & pd.concat(change_conditions, axis=1).any(axis=1)
    ]

    def format_changed_rows(changed_rows):
        changes_list = []
        for _, row in changed_rows.iterrows():
            changes = {'Rule Name': row['Rule Name']}
            for col in common_cols:
                if row[f'{col}_before'] != row[f'{col}_after']:
                    changes[f'{col}_before'] = row[f'{col}_before']
                    changes[f'{col}_after'] = row[f'{col}_after']
            changes_list.append(changes)
        return pd.DataFrame(changes_list)

    changed_df = format_changed_rows(changed)

    summary_data = {
        'Category': ['Added Rules', 'Removed Rules', 'Changed Rules'],
        'Count': [len(added_df), len(removed_df), len(changed_df)]
    }
    summary_df = pd.DataFrame(summary_data)

    return summary_df, added_df, removed_df, changed_df


def apply_excel_formatting(results_dir, base_file_name):
    """
    Applies formatting to an Excel file located in the results directory.
    The first row is styled as the header with a gray background and bold text.
    Column widths are auto-adjusted but limited to 350 pixels.

    :param results_dir: str, path to the results directory
    :param base_file_name: str, name of the Excel file
    :return: str, formatted file path
    """
    file_path = os.path.join(results_dir, base_file_name)

    workbook = load_workbook(file_path)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_font = Font(bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column
            col_letter = get_column_letter(column)

            for cell in column_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width

    workbook.save(file_path)
    return base_file_name