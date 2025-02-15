import os
import pandas as pd
import logging
import re
import shutil
import tempfile
from pathlib import Path
from contextlib import contextmanager
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

class PolicyStatus:
    """정책 상태 관련 상수"""
    AUTO_EXTENSION = [98, 99]  # 자동연장 상태 코드
    DEFAULT_DATE = '19000101'  # 기본 날짜값 (1900년 1월 1일)
    RECENT_PERIOD_DAYS = 90  # 최근 정책 판단 기준 일수 (3개월)

class PolicyType:
    """정책 유형 관련 상수"""
    GROUP = 'GROUP'    # 그룹 정책 (P로 시작)
    NORMAL = 'NORMAL'  # 일반 정책 (F로 시작)
    SERVER = 'SERVER'  # 서버 정책 (S로 시작)
    PAM = 'PAM'       # PAM 정책 (M으로 시작)
    OLD = 'OLD'       # 이전 정책
    UNKNOWN = 'Unknown'  # 알 수 없는 정책

class ExceptionType:
    """예외 정책 유형 관련 상수"""
    EXCEPTION = '예외신청정책'
    AUTO_EXTENSION = '자동연장정책'
    NEW = '신규정책'
    INFRASTRUCTURE = '인프라정책'
    TEST_GROUP = 'test_group_정책'
    DISABLED = '비활성화정책'
    STANDARD = '기준정책'
    BLOCK = '차단정책'

class ExpirationStatus:
    """만료 상태 관련 상수"""
    EXPIRED = '만료'
    NOT_EXPIRED = '미만료'

# Config
COLUMNS = [
    'Rule Name', 'Source', 'User', 'Destination', 'Service', 'Application', 'Description',
    'REQUEST_ID', 'REQUEST_START_DATE', 'REQUEST_END_DATE', 'TITLE', 'REQUESTER_ID',
    'REQUESTER_EMAIL', 'REQUESTER_NAME', 'REQUESTER_DEPT', 'WRITER_PERSON_ID', 'WRITE_PERSON_EMAIL',
    'WRITE_PERSON_NAME', 'WRITE_PERSON_DEPT', 'APPROVAL_PERSON_ID', 'APPROVAL_PERSON_EMAIL',
    'APPROVAL_PERSON_NAME', 'APPROVAL_PERSON_DEPT_NAME'
]

COLUMNS_NO_HISTORY = [
    'Rule Name', 'Source', 'User', 'Destination', 'Service', 'Application', 'Description'
]

DATE_COLUMNS = ['REQUEST_START_DATE', 'REQUEST_END_DATE']

TRANSLATED_COLUMNS = {
    'REQUEST_ID': '신청번호',
    'REQUEST_START_DATE': '시작일',
    'REQUEST_END_DATE': '종료일',
    'TITLE': '제목',
    'REQUESTER_ID': '신청자 ID',
    'REQUESTER_EMAIL': '신청자 이메일',
    'REQUESTER_NAME': '신청자명',
    'REQUESTER_DEPT': '신청자 부서',
    'WRITER_PERSON_ID': '기안자 ID',
    'WRITE_PERSON_EMAIL': '기안자 이메일',
    'WRITE_PERSON_NAME': '기안자명',
    'WRITE_PERSON_DEPT': '기안자 부서',
    'APPROVAL_PERSON_ID': '결재자ID',
    'APPROVAL_PERSON_EMAIL': '결재자 이메일',
    'APPROVAL_PERSON_NAME': '결재자명',
    'APPROVAL_PERSON_DEPT_NAME': '결재자 부서'
}

# 예외 처리 관련 상수
EXCEPT_LIST = [
    'test',
    'sample',
]

class PolicyProcessor:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self._backup_dir = Path("backups")
        self._backup_dir.mkdir(exist_ok=True)

    def _create_backup(self, file_path: str) -> Path:
        """
        파일의 백업을 생성합니다.
        
        Args:
            file_path (str): 백업할 파일 경로
            
        Returns:
            Path: 백업 파일 경로
        """
        source_path = Path(file_path)
        if not source_path.exists():
            raise FileNotFoundError(f"백업할 파일을 찾을 수 없습니다: {file_path}")
            
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self._backup_dir / f"{source_path.stem}_{timestamp}{source_path.suffix}"
        
        shutil.copy2(source_path, backup_path)
        self.logger.info(f"파일 백업 생성 완료: {backup_path}")
        return backup_path
        
    @contextmanager
    def _safe_file_operation(self, file_path: str, mode: str = "wb"):
        """
        안전한 파일 작업을 위한 컨텍스트 매니저
        
        Args:
            file_path (str): 대상 파일 경로
            mode (str): 파일 열기 모드
        """
        target_path = Path(file_path)
        
        # 파일 권한 검사
        if target_path.exists():
            if not os.access(target_path, os.W_OK):
                raise PermissionError(f"파일에 쓰기 권한이 없습니다: {file_path}")
        elif not os.access(target_path.parent, os.W_OK):
            raise PermissionError(f"디렉토리에 쓰기 권한이 없습니다: {target_path.parent}")
            
        # 임시 파일 생성
        temp_fd, temp_path = tempfile.mkstemp(dir=target_path.parent)
        os.close(temp_fd)
        temp_path = Path(temp_path)
        
        try:
            # 백업 생성
            if target_path.exists():
                self._create_backup(file_path)
                
            yield temp_path
            
            # 임시 파일을 대상 파일로 이동 (원자적 작업)
            temp_path.replace(target_path)
            self.logger.debug(f"파일 작업 완료: {file_path}")
            
        except Exception as e:
            if temp_path.exists():
                temp_path.unlink()
            raise e
            
    def _safe_excel_save(self, df: pd.DataFrame, file_path: str, **kwargs):
        """
        데이터프레임을 안전하게 엑셀 파일로 저장합니다.
        
        Args:
            df (pd.DataFrame): 저장할 데이터프레임
            file_path (str): 저장할 파일 경로
            **kwargs: pandas.DataFrame.to_excel에 전달할 추가 인자
        """
        with self._safe_file_operation(file_path) as temp_path:
            df.to_excel(temp_path, **kwargs)
            
    def _safe_workbook_save(self, wb, file_path: str):
        """
        워크북을 안전하게 저장합니다.
        
        Args:
            wb: openpyxl Workbook 객체
            file_path (str): 저장할 파일 경로
        """
        with self._safe_file_operation(file_path) as temp_path:
            wb.save(temp_path)

    def _update_version(self, filename: str, final_version: bool = False) -> str:
        """
        파일 버전을 업데이트하는 내부 메서드
        
        Args:
            filename (str): 원본 파일명
            final_version (bool): 최종 버전 여부
            
        Returns:
            str: 업데이트된 파일명
        """
        base_name, ext = filename.rsplit('.', 1)

        match = re.search(r'_v(\d+)$', base_name)
        final_match = re.search(r'_vf$', base_name)

        if final_match:
            return filename
        
        if final_version:
            if match:
                new_base_name = re.sub(r'_v\d+$', '_vf', base_name)
            else:
                new_base_name = f"{base_name}_vf"
        else:
            if match:
                version = int(match.group(1))
                new_version = version + 1
                new_base_name = re.sub(r'_v\d+$', f'_v{new_version}', base_name)
            else:
                new_base_name = f"{base_name}_v1"
        
        new_filename = f"{new_base_name}.{ext}"
        return new_filename

    def _save_to_excel(self, df: pd.DataFrame, sheet_type: str, file_name: str):
        """
        엑셀 파일 저장 및 스타일 적용하는 내부 메서드
        
        Args:
            df (pd.DataFrame): 저장할 데이터프레임
            sheet_type (str): 시트 유형
            file_name (str): 저장할 파일명
        """
        wb = load_workbook(file_name)
        sheet = wb[sheet_type]

        sheet.insert_rows(1)
        sheet['A1'] = '="대상 정책 수: "&COUNTA(B:B)-1'

        sheet['A1'].font = Font(bold=True)

        for col in range(1, 8):
            cell = sheet.cell(row=2, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        if sheet_type != '이력없음_미사용정책':
            for col in range(8, 24):
                cell = sheet.cell(row=2, column=col)
                cell.fill = PatternFill(start_color='ccffff', end_color='ccffff', fill_type='solid')
        
        self._safe_workbook_save(wb, file_name)

    def parse_request_type(self, file_path: str):
        """
        Description에서 신청번호를 파싱하여 정책 파일을 업데이트합니다.
        
        Args:
            file_path (str): 처리할 정책 파일 경로
            
        Raises:
            FileNotFoundError: 파일을 찾을 수 없는 경우
            pd.errors.EmptyDataError: 빈 파일인 경우
        """
        def convert_to_date(date_str):
            try:
                date_obj = datetime.strptime(date_str, '%Y%m%d')
                return date_obj.strftime('%Y-%m-%d')
            except ValueError:
                return date_str
        
        def parse_request_info(rulename, description):
            data_dict = {
                "Request Type": PolicyType.UNKNOWN,
                "Request ID": None,
                "Ruleset ID": None,
                "MIS ID": None,
                "Request User": None,
                "Start Date": convert_to_date(PolicyStatus.DEFAULT_DATE),
                "End Date": convert_to_date(PolicyStatus.DEFAULT_DATE),
            }

            if pd.isnull(description):
                return data_dict
            
            # initial pattern
            pattern_3 = re.compile("MASKED")
            pattern_1_rulename = re.compile("MASKED")
            pattern_1_user = r'MASKED'
            rulename_1_rulename = r'MASKED'
            rulename_1_date = r'MASKED'

            # matched
            match_3 = pattern_3.match(description)
            name_match = pattern_1_rulename.match(str(rulename))
            user_match = re.search(pattern_1_user, description)
            desc_match = re.search(rulename_1_rulename, description)
            date_match = re.search(rulename_1_date, description)

            if match_3:
                data_dict.update({
                    "Request Type": None,
                    "Request ID": match_3.group(5),
                    "Ruleset ID": match_3.group(1),
                    "MIS ID": match_3.group(6) if match_3.group(6) else None,
                    "Request User": match_3.group(4),
                    "Start Date": convert_to_date(match_3.group(2)),
                    "End Date": convert_to_date(match_3.group(3)),
                })
                
                type_code = data_dict["Request ID"][:1]
                if type_code == "P":
                    data_dict["Request Type"] = PolicyType.GROUP
                elif type_code == "F":
                    data_dict["Request Type"] = PolicyType.NORMAL
                elif type_code == "S":
                    data_dict["Request Type"] = PolicyType.SERVER
                elif type_code == "M":
                    data_dict["Request Type"] = PolicyType.PAM
                else:
                    data_dict["Request Type"] = PolicyType.UNKNOWN
            
            if name_match:
                data_dict.update({
                    'Request Type': PolicyType.OLD,
                    'Request ID': name_match.group(1)
                })
                if user_match:
                    data_dict['Request User'] = user_match.group(1).replace("*ACL*", "")
                if date_match:
                    data_dict['Start Date'] = convert_to_date(date_match.group().split("~")[0])
                    data_dict['End Date'] = convert_to_date(date_match.group().split("~")[1])
            
            if desc_match:
                date = description.split(';')[0]
                start_date = date.split('~')[0].replace('[', '').replace('-', '')
                end_date = date.split('~')[1].replace(']', '').replace('-', '')

                data_dict.update({
                    "Request Type": PolicyType.OLD,
                    "Request ID": desc_match.group(1).split('-')[1],
                    "Ruleset ID": None,
                    "MIS ID": None,
                    "Request User": user_match.group(1).replace("*ACL*", "") if user_match else None,
                    "Start Date": convert_to_date(start_date),
                    "End Date": convert_to_date(end_date),
                })
            
            return data_dict
        
        try:
            df = pd.read_excel(file_path)
            if df.empty:
                self.logger.error(f"빈 파일입니다: {file_path}")
                raise pd.errors.EmptyDataError("빈 파일입니다")
                
            total = len(df)
            self.logger.info(f"정책 파싱 시작: 총 {total}개 처리 예정")
            
            for index, row in df.iterrows():
                if (index + 1) % 100 == 0:
                    self.logger.debug(f"진행률: {((index + 1)/total)*100:.1f}% ({index + 1}/{total})")
                    
                result = parse_request_info(row['Rule Name'], row['Description'])
                for key, value in result.items():
                    df.at[index, key] = value
            
            self.logger.info("정책 파싱 완료")
            output_file = self._update_version(file_path)
            self._safe_excel_save(df, output_file, index=False)
            
        except FileNotFoundError:
            self.logger.error(f"파일을 찾을 수 없습니다: {file_path}")
            raise
        except Exception as e:
            self.logger.error(f"정책 파싱 중 오류 발생: {e}")
            raise

    def extract_request_id(self, file_path: str):
        """
        정책 파일에서 신청번호를 추출합니다.
        
        Args:
            file_path (str): 처리할 정책 파일 경로
            
        Raises:
            FileNotFoundError: 파일을 찾을 수 없는 경우
            pd.errors.EmptyDataError: 빈 파일인 경우
        """
        try:
            df = pd.read_excel(file_path)
            if df.empty:
                self.logger.error(f"빈 파일입니다: {file_path}")
                raise pd.errors.EmptyDataError("빈 파일입니다")

            unique_types = df[df['Request Type'] != 'Unknown']['Request Type'].unique()
            selected_types = unique_types[:5]
            selected_data = df[df['Request Type'].isin(selected_types)]

            output_file = f"request_id_{file_path}"
            with self._safe_file_operation(output_file) as temp_path:
                with pd.ExcelWriter(temp_path) as writer:
                    for request_type, group in selected_data.groupby('Request Type'):
                        group[['Request ID']].drop_duplicates().to_excel(
                            writer, 
                            sheet_name=request_type, 
                            index=False
                        )
            self.logger.info(f"신청번호 추출 완료: {output_file}")
            
        except FileNotFoundError:
            self.logger.error(f"파일을 찾을 수 없습니다: {file_path}")
            raise
        except Exception as e:
            self.logger.error(f"신청번호 추출 중 오류 발생: {e}")
            raise

    def add_request_info(self, rule_file_path: str, info_file_path: str, progress_callback=None):
        """
        정책 파일에 신청 정보를 추가합니다.
        
        Args:
            rule_file_path (str): 처리할 정책 파일 경로
            info_file_path (str): 신청 정보 파일 경로
            progress_callback (callable, optional): 진행률을 전달받을 콜백 함수
            
        Raises:
            FileNotFoundError: 파일을 찾을 수 없는 경우
            pd.errors.EmptyDataError: 빈 파일인 경우
        """
        def read_and_process_excel(file):
            try:
                df = pd.read_excel(file)
                if df.empty:
                    raise pd.errors.EmptyDataError(f"빈 파일입니다: {file}")
                df.replace({'nan': None}, inplace=True)
                return df.astype(str)
            except FileNotFoundError:
                self.logger.error(f"파일을 찾을 수 없습니다: {file}")
                raise
            except Exception as e:
                self.logger.error(f"파일 처리 중 오류 발생: {e}")
                raise
        
        def match_and_update_df(rule_df, info_df):
            total = len(rule_df)
            self.logger.info(f"정책 정보 매칭 시작: 총 {total}개 처리 예정")
            
            for idx, row in rule_df.iterrows():
                # 진행률 업데이트
                progress = ((idx + 1) / total) * 100
                if progress_callback:
                    progress_callback(progress)
                elif (idx + 1) % 100 == 0:
                    self.logger.debug(f"진행률: {progress:.1f}% ({idx + 1}/{total})")
                
                if row['Request Type'] == 'GROUP':
                    matched_row = info_df[
                        ((info_df['REQUEST_ID'] == row['Request ID']) & (info_df['MIS_ID'] == row['MIS ID'])) |
                        ((info_df['REQUEST_ID'] == row['Request ID']) & (info_df['REQUEST_END_DATE'] == row['End Date']) & (info_df['WRITE_PERSON_ID'] == row['Request User'])) |
                        ((info_df['REQUEST_ID'] == row['Request ID']) & (info_df['REQUEST_END_DATE'] == row['End Date']) & (info_df['REQUESTER_ID'] == row['Request User']))
                    ]
                else:
                    matched_row = info_df[info_df['REQUEST_ID'] == row['Request ID']]
                
                if not matched_row.empty:
                    for col in matched_row.columns:
                        if col in ['REQUEST_START_DATE', 'REQUEST_END_DATE', 'Start Date', 'End Date']:
                            rule_df.at[idx, col] = pd.to_datetime(matched_row[col].values[0], errors='coerce')
                        else:
                            rule_df.at[idx, col] = matched_row[col].values[0]
                elif row['Request Type'] not in ['nan', 'Unknown']:
                    rule_df.at[idx, 'REQUEST_ID'] = row['Request ID']
                    rule_df.at[idx, 'REQUEST_START_DATE'] = row['Start Date']
                    rule_df.at[idx, 'REQUEST_END_DATE'] = row['End Date']
                    rule_df.at[idx, 'REQUESTER_ID'] = row['Request User']
                    rule_df.at[idx, 'REQUESTER_EMAIL'] = row['Request User'] + '@gmail.com'
            
            self.logger.info("정책 정보 매칭 완료")

        try:
            rule_df = read_and_process_excel(rule_file_path)
            info_df = read_and_process_excel(info_file_path)
            
            info_df = info_df.sort_values(by='REQUEST_END_DATE', ascending=False)
            auto_extension_id = info_df[info_df['REQUEST_STATUS'].isin(PolicyStatus.AUTO_EXTENSION)]['REQUEST_ID'].drop_duplicates()
            
            match_and_update_df(rule_df, info_df)
            rule_df.replace({'nan': None}, inplace=True)
            rule_df.loc[rule_df['REQUEST_ID'].isin(auto_extension_id), 'REQUEST_STATUS'] = '99'
            
            output_file = self._update_version(rule_file_path)
            self._safe_excel_save(rule_df, output_file, index=False)
            self.logger.info(f"정책 정보 업데이트 완료: {output_file}")
            
        except Exception as e:
            self.logger.error(f"정책 정보 추가 중 오류 발생: {e}")
            raise

    def process_exceptions(self, file_path: str, vendor_type: str = 'paloalto'):
        """
        정책 파일에서 예외 처리를 수행합니다.
        
        Args:
            file_path (str): 처리할 정책 파일 경로
            vendor_type (str): 벤더 타입 ('paloalto' 또는 'secui')
            
        Raises:
            FileNotFoundError: 파일을 찾을 수 없는 경우
            ValueError: 지원하지 않는 벤더 타입이 지정된 경우
        """
        if vendor_type not in ['paloalto', 'secui']:
            raise ValueError("지원하지 않는 벤더 타입입니다. 'paloalto' 또는 'secui'만 사용 가능합니다.")

        try:
            df = pd.read_excel(file_path)
        except FileNotFoundError:
            self.logger.error(f"파일을 찾을 수 없습니다: {file_path}")
            raise
        except Exception as e:
            self.logger.error(f"엑셀 파일 처리 중 오류 발생: {e}")
            raise

        current_date = datetime.now()
        three_months_ago = current_date - timedelta(days=PolicyStatus.RECENT_PERIOD_DAYS)

        df["예외"] = ''

        # 1. 예외 신청정책 처리
        df['REQUEST_ID'] = df['REQUEST_ID'].fillna('')
        for id in EXCEPT_LIST:
            df.loc[df['REQUEST_ID'].str.startswith(id, na=False), '예외'] = ExceptionType.EXCEPTION
        
        # 2. 자동연장 정책 처리
        df.loc[df['REQUEST_STATUS'].isin(PolicyStatus.AUTO_EXTENSION), '예외'] = ExceptionType.AUTO_EXTENSION

        # 3. 신규 정책 처리
        df['날짜'] = df['Rule Name'].str.extract(r'(\d{8})', expand=False)
        df['날짜'] = pd.to_datetime(df['날짜'], format='%Y%m%d', errors='coerce')
        df.loc[(df['날짜'] >= three_months_ago) & (df['날짜'] <= current_date), '예외'] = ExceptionType.NEW

        if vendor_type == 'paloalto':
            # 4. 인프라 정책 처리
            deny_std_rule_index = df[df['Rule Name'] == 'deny_rule'].index[0]
            df.loc[df.index < deny_std_rule_index, '예외'] = ExceptionType.INFRASTRUCTURE

            # 5. 테스트 그룹 정책 처리
            df.loc[df['Rule Name'].str.startswith(('sample_', 'test_')), '예외'] = ExceptionType.TEST_GROUP

            # 6. 비활성화 정책 처리
            df.loc[df['Enable'] == 'N', '예외'] = ExceptionType.DISABLED

            # 7. 기준 정책 처리
            df.loc[(df['Rule Name'].str.endswith('_Rule')) & (df['Enable'] == 'N'), '예외'] = ExceptionType.STANDARD

        elif vendor_type == 'secui':
            # 4. 인프라 정책 처리
            deny_std_rule_index = df[df['Description'].str.contains('deny_rule')].index[0]
            df.loc[df.index < deny_std_rule_index, '예외'] = ExceptionType.INFRASTRUCTURE

            # 5. 테스트 그룹 정책 처리
            df.loc[df['Description'].str.contains(('sample_', 'test_')), '예외'] = ExceptionType.TEST_GROUP

            # 6. 비활성화 정책 처리
            df.loc[df['Enable'] == 'N', '예외'] = ExceptionType.DISABLED

            # 7. 기준 정책 처리
            df.loc[(df['Description'].str.contains('기준룰')) & (df['Enable'] == 'N'), '예외'] = ExceptionType.STANDARD

        # 8. 차단 정책 처리
        df.loc[df['Action'] == 'deny', '예외'] = ExceptionType.BLOCK

        df['예외'].fillna('', inplace=True)

        # 컬럼 순서 정리
        cols = list(df.columns)
        cols = ['예외'] + [col for col in cols if col != '예외']
        df = df[cols]

        # 만료여부 체크
        df['만료여부'] = df.apply(
            lambda row: ExpirationStatus.NOT_EXPIRED 
            if pd.to_datetime(row['REQUEST_END_DATE']) > datetime.now() 
            else ExpirationStatus.EXPIRED 
            if pd.notna(row['REQUEST_END_DATE']) 
            else ExpirationStatus.EXPIRED, 
            axis=1
        )

        df.drop(columns=['날짜'], inplace=True)
        df.rename(columns={'Request Type': '신청이력'}, inplace=True)
        df.drop(columns=['Request ID', 'Ruleset ID', 'MIS ID', 'Request User', 'Start Date', 'End Date'], inplace=True)

        # 최종 컬럼 순서 조정
        cols = list(df.columns)
        cols.insert(cols.index('예외') + 1, cols.pop(cols.index('만료여부')))
        df = df[cols]
        cols.insert(cols.index('예외') + 1, cols.pop(cols.index('신청이력')))
        df = df[cols]
        cols.insert(cols.index('만료여부') + 1, '미사용여부')
        df['미사용여부'] = ''

        output_file = self._update_version(file_path, True)
        self._safe_excel_save(df, output_file, index=False)

    def organize_redundant_file(self, file_path: str):
        """
        중복 정책을 분류하고 처리합니다.
        
        Args:
            file_path (str): 처리할 정책 파일 경로
        """
        expected_columns = ['No', 'Type', 'Seq', 'Rule Name', 'Enable', 'Action', 'Source', 'User', 'Destination', 'Service', 'Application', 'Security Profile', 'Description', 'Request Type', 'Request ID', 'Ruleset ID', 'MIS ID', 'Request User', 'Start Date', 'End Date']
        expected_columns_2 = ['No', 'Type', 'Vsys', 'Seq', 'Rule Name', 'Enable', 'Action', 'Source', 'User', 'Destination', 'Service', 'Application', 'Security Profile', 'Description', 'Request Type', 'Request ID', 'Ruleset ID', 'MIS ID', 'Request User', 'Start Date', 'End Date']

        try:
            df = pd.read_excel(file_path)
            current_columns = df.columns.tolist()

            if current_columns != expected_columns and current_columns != expected_columns_2:
                self.logger.error('컬럼명이 일치하지 않습니다.')
                return
        except Exception as e:
            self.logger.error(f'엑셀 파일을 열 수 없습니다: {e}')
            return

        # 자동연장 정책 확인
        auto_extension_id = df[df['REQUEST_STATUS'].isin([98, 99])]['REQUEST_ID'].drop_duplicates()
        df['자동연장'] = df['Request ID'].isin(auto_extension_id)

        # 종료일 기준 검증
        df['늦은종료일'] = df.groupby('No')['End Date'].transform(lambda x: (x == x.max()) & (~x.duplicated(keep='first')))
        df['신청자검증'] = df.groupby('No')['Request User'].transform(lambda x: x.nunique() == 1)

        # 상위 정책 검증
        target_rule_true = df[(df['Type'] == 'Upper') & (df['늦은종료일'] == True)]['No'].unique()
        df['날짜검증'] = False
        df.loc[df['No'].isin(target_rule_true), '날짜검증'] = True

        # 작업 구분
        df['작업구분'] = '유지'
        df.loc[df['늦은종료일'] == False, '작업구분'] = '삭제'

        # 공지 여부
        df['공지여부'] = False
        df.loc[df['신청자검증'] == False, '공지여부'] = True

        # 미사용 예외
        df['미사용예외'] = False
        df.loc[(df['날짜검증'] == False) & (df['늦은종료일'] == True), '미사용예외'] = True

        # 자동연장 정책 처리
        extensioned_df = df.groupby('No').filter(lambda x: x['자동연장'].any())
        extensioned_group = extensioned_df[extensioned_df['Request Type'] == 'GROUP']
        exception_target = extensioned_group.groupby('No').filter(lambda x: len(x['Request ID'].unique()) >= 2)
        exception_id = exception_target[(exception_target['자동연장'] == True) & (exception_target['작업구분'] == '삭제')]['No']

        df = df[~df['No'].isin(exception_id)]

        # 필터링
        filtered_no = df.groupby('No').filter(
            lambda x: (x['Request Type'] != 'GROUP').any() and
                    (x['작업구분'] == '삭제').any() and
                    (x['자동연장'] == True).any()
        )['No'].unique()
        df = df[~df['No'].isin(filtered_no)]

        filtered_no_2 = df.groupby('No').filter(
            lambda x: (x['작업구분'] != '유지').all()
        )['No'].unique()
        df = df[~df['No'].isin(filtered_no_2)]

        # 결과 파일 생성
        notice_df = df[df['공지여부'] == True].copy()
        delete_df = df[df['공지여부'] == False].copy()

        # 컬럼 정리
        for result_df in [notice_df, delete_df]:
            column_to_move = result_df.pop('작업구분')
            result_df.insert(0, '작업구분', column_to_move)
            result_df.drop(['Request Type', 'Ruleset ID', 'MIS ID', 'Start Date', 'End Date', 
                          '늦은종료일', '신청자검증', '날짜검증', '공지여부', '미사용예외', '자동연장'], 
                          axis=1, inplace=True)

        # 파일 저장
        base_name = os.path.splitext(file_path)[0]
        self._safe_excel_save(df, f'{base_name}_정리.xlsx', index=False)
        self._safe_excel_save(notice_df, f'{base_name}_공지.xlsx', index=False)
        self._safe_excel_save(delete_df, f'{base_name}_삭제.xlsx', index=False)

    def add_mis_id(self, rule_file_path: str, mis_file_path: str):
        """
        정책 파일에 MIS ID 정보를 추가합니다.
        
        Args:
            rule_file_path (str): 처리할 정책 파일 경로
            mis_file_path (str): MIS ID 정보가 있는 CSV 파일 경로
            
        Raises:
            FileNotFoundError: 파일을 찾을 수 없는 경우
            pd.errors.EmptyDataError: 빈 파일인 경우
        """
        try:
            mis_df = pd.read_csv(mis_file_path)
            rule_df = pd.read_excel(rule_file_path)
            
            if mis_df.empty or rule_df.empty:
                raise pd.errors.EmptyDataError("빈 파일이 있습니다")

            self.logger.info(f"MIS ID 정보 업데이트 시작: 정책 {len(rule_df)}개, MIS ID {len(mis_df)}개")
            
            mis_df_unique = mis_df.drop_duplicates(subset=['ruleset_id'], keep='first')
            mis_id_map = mis_df_unique.set_index('ruleset_id')['mis_id']

            rule_df['MIS ID'] = rule_df.apply(
                lambda row: mis_id_map.get(row['Ruleset ID'], row['MIS ID']) 
                if pd.isna(row['MIS ID']) or row['MIS ID'] == '' 
                else row['MIS ID'], 
                axis=1
            )

            output_file = self._update_version(rule_file_path)
            self._safe_excel_save(rule_df, output_file, index=False)
            self.logger.info(f"MIS ID 정보 업데이트 완료: {output_file}")
            
        except FileNotFoundError as e:
            self.logger.error(f"파일을 찾을 수 없습니다: {e.filename}")
            raise
        except Exception as e:
            self.logger.error(f"MIS ID 정보 추가 중 오류 발생: {e}")
            raise

    def notice_file_organization(self, file_path: str):
        """
        정책 파일을 분류하여 공지용 파일들을 생성합니다.
        
        Args:
            file_path (str): 처리할 정책 파일 경로
        """
        try:
            df = pd.read_excel(file_path)
            base_name = os.path.splitext(file_path)[0]

            # 만료된 사용 정책
            filtered_df = df[
                ((df['예외'].isna()) | (df['예외'] == '신규정책')) &
                (df['중복여부'].isna()) &
                (df['신청이력'] != 'Unknown') &
                (df['만료여부'] == '만료') &
                (df['미사용여부'] == '사용')
            ]
            if not filtered_df.empty:
                self._save_notice_file(filtered_df, f'{base_name}_기간만료(공지용).xlsx', '만료_사용정책')

            # 만료된 미사용 정책
            filtered_df = df[
                ((df['예외'].isna()) | (df['예외'] == '신규정책')) &
                (df['중복여부'].isna()) &
                (df['신청이력'] != 'Unknown') &
                (df['만료여부'] == '만료') &
                (df['미사용여부'] == '미사용')
            ]
            if not filtered_df.empty:
                self._save_notice_file(filtered_df, f'{base_name}_만료_미사용정책(공지용).xlsx', '만료_미사용정책')

            # 장기 미사용 정책
            filtered_df = df[
                (df['예외'].isna()) &
                (df['중복여부'].isna()) &
                (df['신청이력'].isin(['GROUP', 'NORMAL', 'OLD'])) &
                (df['만료여부'] == '미만료') &
                (df['미사용여부'] == '미사용')
            ]
            if not filtered_df.empty:
                self._save_notice_file(filtered_df, f'{base_name}_장기미사용정책(공지용).xlsx', '미만료_미사용정책')

            # 이력 없는 미사용 정책
            filtered_df = df[
                (df['예외'].isna()) &
                (df['중복여부'].isna()) &
                (df['신청이력'] != 'Unknown') &
                (df['만료여부'] == '만료') &
                (df['미사용여부'] == '미사용')
            ]
            if not filtered_df.empty:
                self._save_notice_file(filtered_df, f'{base_name}_이력없는_미사용정책.xlsx', '이력없음_미사용정책')

        except Exception as e:
            self.logger.error(f"정책 분류 중 오류 발생: {e}")

    def _save_notice_file(self, df: pd.DataFrame, filename: str, sheet_type: str):
        """공지용 파일을 저장하는 내부 메서드"""
        selected_df = df[COLUMNS].copy()
        selected_df = selected_df.astype(str)

        for date_column in DATE_COLUMNS:
            selected_df[date_column] = pd.to_datetime(selected_df[date_column]).dt.strftime('%Y-%m-%d')
        
        selected_df.rename(columns=TRANSLATED_COLUMNS, inplace=True)
        selected_df.fillna('', inplace=True)
        selected_df.replace('nan', '', inplace=True)
        
        self._safe_excel_save(selected_df, filename, index=False, na_rep='', sheet_name=sheet_type)
        self._save_to_excel(selected_df, sheet_type, filename)