import logging
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class PolicyAnalyzer:
    """정책 분석을 담당하는 클래스"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def analyze_duplicates(self, policies_df, vendor, output_dir):
        """중복 정책 분석 수행
        
        Args:
            policies_df (pd.DataFrame): 분석할 정책 데이터프레임
            vendor (str): 방화벽 벤더 ('paloalto' or 'secui')
            output_dir (str): 결과 파일을 저장할 디렉토리 경로
        
        Returns:
            str: 생성된 분석 파일의 경로
        """
        try:
            self.logger.info("중복 정책 분석 시작")
            
            # 활성화된 허용 정책만 필터링
            df_filtered = policies_df[
                (policies_df['Enable'] == 'Y') & 
                (policies_df['Action'] == 'Allow')
            ]

            # 비교할 컬럼 설정
            columns_to_check = [
                'Enable', 'Action', 'Source', 'User', 
                'Destination', 'Service', 'Application'
            ]
            
            # 벤더별 추가 설정
            if 'Vsys' in policies_df.columns:
                columns_to_check.append('Vsys')
            
            if vendor == 'paloalto':
                df_filtered['Service'] = df_filtered['Service'].str.replace('_', '-')
                columns_to_check.append('Category')
            
            df_check = df_filtered[columns_to_check]

            # 정책 정규화 및 중복 체크
            def normalize_policy(policy_series):
                normalized_policy = policy_series.apply(
                    lambda x: ','.join(sorted(str(x).split(','))) if pd.notna(x) else ''
                )
                return tuple(normalized_policy)
            
            policy_map = {}
            results_list = []
            current_no = 1

            self.logger.info('중복 정책 확인 중...')
            
            for idx, row in df_filtered.iterrows():
                try:
                    current_policy = normalize_policy(df_check.iloc[idx])
                    if current_policy in policy_map:
                        row_dict = row.to_dict()
                        row_dict.update({'No': policy_map[current_policy], 'Type': 'Lower'})
                        results_list.append(row_dict)
                    else:
                        policy_map[current_policy] = current_no
                        row_dict = row.to_dict()
                        row_dict.update({'No': current_no, 'Type': 'Upper'})
                        results_list.append(row_dict)
                        current_no += 1
                except Exception as e:
                    self.logger.error(f'정책 확인 중 오류 발생 (인덱스 {idx}): {str(e)}')
                    continue

            # 결과를 DataFrame으로 변환
            results = pd.DataFrame(results_list)

            # Upper와 Lower가 모두 있는 그룹만 필터링
            def ensure_upper_and_lower(df):
                valid_groups = []
                for _, group in df.groupby('No'):
                    if set(group['Type']) == {'Upper', 'Lower'}:
                        valid_groups.append(group)
                return pd.concat(valid_groups) if valid_groups else pd.DataFrame()

            duplicated_results = ensure_upper_and_lower(results)
            
            if duplicated_results.empty:
                self.logger.info("중복된 정책이 없습니다.")
                return None

            # 그룹 번호 재할당
            duplicated_results['No'] = duplicated_results.groupby('No').ngroup() + 1

            # 컬럼 순서 정렬
            columns_order = ['No', 'Type'] + [col for col in policies_df.columns]
            duplicated_results = duplicated_results[columns_order]
            duplicated_results = duplicated_results.sort_values(
                by=['No', 'Type'], 
                ascending=[True, False]
            )

            # 결과 파일 생성
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = Path(output_dir) / f'duplicate_analysis_{timestamp}.xlsx'

            # 엑셀 스타일 설정
            upper_fill = PatternFill(
                start_color="daeef3", 
                end_color="daeef3", 
                fill_type="solid"
            )
            lower_fill = PatternFill(
                start_color="f2f2f2", 
                end_color="f2f2f2", 
                fill_type="solid"
            )
            header_fill = PatternFill(
                start_color="00b0f0", 
                end_color="00b0f0", 
                fill_type="solid"
            )
            header_font = Font(bold=True, color='FFFFFF')

            self.logger.info("결과 저장 중...")
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                if 'Vsys' in policies_df.columns:
                    # Vsys별로 시트 생성
                    for vsys, vsys_df in duplicated_results.groupby('Vsys'):
                        sheet_name = f'Analysis_{vsys}'
                        vsys_df.to_excel(writer, index=False, sheet_name=sheet_name)
                        worksheet = writer.sheets[sheet_name]

                        # 헤더 스타일 적용
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                        
                        # 데이터 스타일 적용
                        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                            for cell in row:
                                if row[1].value == 'Upper':
                                    cell.fill = upper_fill
                                elif row[1].value == 'Lower':
                                    cell.fill = lower_fill
                else:
                    # 단일 시트 생성
                    duplicated_results.to_excel(writer, index=False, sheet_name='Analysis')
                    worksheet = writer.sheets['Analysis']

                    # 헤더 스타일 적용
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    # 데이터 스타일 적용
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                        for cell in row:
                            if row[1].value == 'Upper':
                                cell.fill = upper_fill
                            elif row[1].value == 'Lower':
                                cell.fill = lower_fill

            self.logger.info(f"분석 결과가 {output_file}에 저장되었습니다.")
            return str(output_file)

        except Exception as e:
            self.logger.error(f"중복 정책 분석 중 오류 발생: {str(e)}")
            raise
    
    def analyze_usage_patterns(self, policies_df, usage_data):
        """정책 사용 패턴 분석
        
        Args:
            policies_df (pd.DataFrame): 정책 데이터프레임
            usage_data (pd.DataFrame): 사용이력 데이터
        
        Returns:
            pd.DataFrame: 사용 패턴이 분석된 데이터프레임
        """
        try:
            # 사용이력 데이터와 정책 데이터 병합
            merged_df = policies_df.merge(
                usage_data,
                how='left',
                left_on='Rule Name',
                right_on='rule_name'
            )
            
            # 미사용 정책 표시
            merged_df['usage_status'] = merged_df['hit_count'].apply(
                lambda x: '미사용' if pd.isna(x) or x == 0 else '사용'
            )
            
            return merged_df
        except Exception as e:
            self.logger.error(f"사용 패턴 분석 중 오류 발생: {str(e)}")
            raise 