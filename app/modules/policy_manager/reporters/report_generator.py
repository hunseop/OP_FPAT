import logging
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class BaseReportGenerator:
    """보고서 생성을 위한 기본 클래스"""

    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir)
        self.logger = logging.getLogger(__name__)
        
        # 기본 스타일 설정
        self.header_style = {
            'font': Font(bold=True),
            'fill': PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center')
        }
        
        self.data_style = {
            'alignment': Alignment(horizontal='left', vertical='center')
        }

    def _save_report(self, df: pd.DataFrame, filename: str, sheet_name: str = 'Sheet1') -> str:
        """데이터프레임을 엑셀 파일로 저장
        
        Args:
            df: 저장할 데이터프레임
            filename: 파일명
            sheet_name: 시트명
            
        Returns:
            저장된 파일의 경로
        """
        try:
            # 출력 디렉토리 생성
            self.output_dir.mkdir(parents=True, exist_ok=True)
            
            # 파일명에 타임스탬프 추가
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_path = self.output_dir / f"{filename}_{timestamp}.xlsx"
            
            # 엑셀 파일로 저장
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 워크시트 가져오기
                worksheet = writer.sheets[sheet_name]
                
                # 헤더 스타일 적용
                for cell in worksheet[1]:
                    cell.font = self.header_style['font']
                    cell.fill = self.header_style['fill']
                    cell.alignment = self.header_style['alignment']
                
                # 데이터 스타일 적용
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = self.data_style['alignment']
                
                # 컬럼 너비 자동 조정
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            return str(file_path)
            
        except Exception as e:
            self.logger.error(f"보고서 저장 중 오류 발생: {str(e)}")
            raise

    def _translate_columns(self, df: pd.DataFrame, translation_map: dict) -> pd.DataFrame:
        """컬럼명 번역
        
        Args:
            df: 변환할 데이터프레임
            translation_map: 번역 매핑 딕셔너리
            
        Returns:
            컬럼명이 변환된 데이터프레임
        """
        try:
            return df.rename(columns=translation_map)
        except Exception as e:
            self.logger.error(f"컬럼명 변환 중 오류 발생: {str(e)}")
            raise

    def _fill_empty_values(self, df: pd.DataFrame, fill_value: str = '-') -> pd.DataFrame:
        """빈 값 채우기
        
        Args:
            df: 처리할 데이터프레임
            fill_value: 채울 값
            
        Returns:
            빈 값이 채워진 데이터프레임
        """
        try:
            return df.fillna(fill_value)
        except Exception as e:
            self.logger.error(f"빈 값 채우기 중 오류 발생: {str(e)}")
            raise 