import pandas as pd
import os

# 템플릿 데이터 생성
template_data = {
    'name': [
        '본사 방화벽',
        '데이터센터 방화벽',
        '클라우드 방화벽',
        '지사 방화벽 1',
        '지사 방화벽 2',
        '개발망 방화벽',
        'DMZ 방화벽',
        '백업센터 방화벽',
        '테스트망 방화벽',
        'VPN 방화벽'
    ],
    'type': [
        'NGF',
        'PALOALTO',
        'PALOALTO',
        'NGF',
        'NGF',
        'MF2',
        'PALOALTO',
        'NGF',
        'MF2',
        'PALOALTO'
    ],
    'ip_address': [
        '192.168.0.1',
        '10.0.0.1',
        '172.16.0.1',
        '192.168.1.1',
        '192.168.2.1',
        '10.10.0.1',
        '172.20.0.1',
        '192.168.10.1',
        '10.20.0.1',
        '172.30.0.1'
    ],
    'username': [
        'admin',
        'admin',
        'admin',
        'admin',
        'admin',
        'admin',
        'admin',
        'admin',
        'admin',
        'admin'
    ],
    'password': [
        'password123',
        'secure456',
        'cloud789',
        'branch123',
        'branch456',
        'dev789',
        'dmz123',
        'backup456',
        'test789',
        'vpn123'
    ]
}

# DataFrame 생성
df = pd.DataFrame(template_data)

# 파일 저장 경로
template_path = 'app/static/templates/firewall_template.xlsx'

# 디렉토리가 없으면 생성
os.makedirs(os.path.dirname(template_path), exist_ok=True)

# 엑셀 파일로 저장
with pd.ExcelWriter(template_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False)
    
    # 워크시트 가져오기
    worksheet = writer.sheets['Sheet1']
    
    # 컬럼 너비 조정
    worksheet.column_dimensions['A'].width = 25  # name
    worksheet.column_dimensions['B'].width = 15  # type
    worksheet.column_dimensions['C'].width = 15  # ip_address
    worksheet.column_dimensions['D'].width = 15  # username
    worksheet.column_dimensions['E'].width = 15  # password

    # 헤더 스타일 설정
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # 데이터 셀 가운데 정렬
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

print(f"템플릿 파일이 생성되었습니다: {template_path}") 