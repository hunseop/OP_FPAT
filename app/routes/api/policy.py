from flask import Blueprint, jsonify, request, send_file, current_app
from app import db
from app.models import SecurityRule, Firewall
import io
import openpyxl

bp = Blueprint('policy_api', __name__)

@bp.route('/list', methods=['GET'])
def get_policies():
    try:
        # 페이지네이션 파라미터
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        
        # 정렬 파라미터
        sort_by = request.args.get('sort_by', 'seq')
        sort_desc = request.args.get('sort_desc', 'false').lower() == 'true'

        # 필터 파라미터 처리
        filters = []
        for key, value in request.args.items():
            if key.startswith('filters['):
                index = int(key.split('[')[1].split(']')[0])
                if len(filters) <= index:
                    filters.append({})
                field = key.split('[')[2].split(']')[0]
                filters[index][field] = value

        # 기본 쿼리 생성
        query = db.session.query(SecurityRule, Firewall.name.label('firewall_name')).join(
            Firewall, SecurityRule.firewall_id == Firewall.id
        )

        # 필터 적용
        for filter_item in filters:
            column = filter_item.get('column')
            operator = filter_item.get('operator')
            value = filter_item.get('value')

            if column and operator and value:
                if column == 'firewall_name':
                    column_obj = Firewall.name
                else:
                    column_obj = getattr(SecurityRule, column, None)

                if column_obj is not None:
                    if operator == 'contains':
                        query = query.filter(column_obj.ilike(f'%{value}%'))
                    elif operator == 'equals':
                        if column == 'enabled':
                            value = value == '활성'
                        query = query.filter(column_obj == value)
                    elif operator == 'starts':
                        query = query.filter(column_obj.ilike(f'{value}%'))
                    elif operator == 'ends':
                        query = query.filter(column_obj.ilike(f'%{value}'))

        # 정렬 적용
        if sort_by == 'firewall_name':
            order_by = Firewall.name.desc() if sort_desc else Firewall.name
        else:
            column = getattr(SecurityRule, sort_by, None)
            if column is not None:
                order_by = column.desc() if sort_desc else column
            else:
                order_by = SecurityRule.seq

        query = query.order_by(order_by)

        # 페이지네이션 적용
        paginated = query.paginate(page=page, per_page=per_page)
        
        # 결과 변환
        policies = []
        for rule, firewall_name in paginated.items:
            policies.append({
                'id': rule.id,
                'firewall_name': firewall_name,
                'seq': rule.seq,
                'name': rule.name,
                'enabled': '활성' if rule.enabled else '비활성',
                'action': '허용' if rule.action == 'allow' else '차단',
                'source': rule.source,
                'user': rule.user,
                'destination': rule.destination,
                'service': rule.service,
                'application': rule.application,
                'security_profile': rule.security_profile,
                'category': rule.category,
                'last_hit': rule.last_hit.strftime('%Y-%m-%d %H:%M:%S') if rule.last_hit else None,
                'description': rule.description
            })

        return jsonify({
            'success': True,
            'data': {
                'policies': policies,
                'total': paginated.total,
                'pages': paginated.pages,
                'current_page': paginated.page,
                'per_page': paginated.per_page
            }
        })

    except Exception as e:
        current_app.logger.error(f"정책 목록 조회 중 오류 발생: {str(e)}")
        return jsonify({
            'success': False,
            'error': '정책 목록을 불러오는 중 오류가 발생했습니다.'
        }), 500

@bp.route('/export', methods=['GET'])
def export_policies():
    try:
        # 쿼리 파라미터 처리
        filters = []
        for key, value in request.args.items():
            if key.startswith('filters['):
                index = int(key.split('[')[1].split(']')[0])
                if len(filters) <= index:
                    filters.append({})
                field = key.split('[')[2].split(']')[0]
                filters[index][field] = value

        # 기본 쿼리 생성
        query = db.session.query(SecurityRule, Firewall.name.label('firewall_name')).join(
            Firewall, SecurityRule.firewall_id == Firewall.id
        )

        # 필터 적용
        for filter_item in filters:
            column = filter_item.get('column')
            operator = filter_item.get('operator')
            value = filter_item.get('value')

            if column and operator and value:
                if column == 'firewall_name':
                    column_obj = Firewall.name
                else:
                    column_obj = getattr(SecurityRule, column, None)

                if column_obj is not None:
                    if operator == 'contains':
                        query = query.filter(column_obj.ilike(f'%{value}%'))
                    elif operator == 'equals':
                        if column == 'enabled':
                            value = value == '활성'
                        query = query.filter(column_obj == value)
                    elif operator == 'starts':
                        query = query.filter(column_obj.ilike(f'{value}%'))
                    elif operator == 'ends':
                        query = query.filter(column_obj.ilike(f'%{value}'))

        # 모든 데이터를 메모리에 로드
        results = query.all()

        # 엑셀 파일 생성
        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = '정책 목록'

        # 헤더 추가
        headers = ['방화벽', '순서', '이름', '상태', '동작', '출발지', '사용자', '목적지', '서비스', 
                  '애플리케이션', '보안 프로필', '카테고리', '마지막 적중', '설명']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # 데이터 행 추가
        for row_idx, (rule, firewall_name) in enumerate(results, 2):
            row_data = [
                firewall_name,
                rule.seq,
                rule.name,
                '활성' if rule.enabled else '비활성',
                '허용' if rule.action == 'allow' else '차단',
                rule.source or '',
                rule.user or '',
                rule.destination or '',
                rule.service or '',
                rule.application or '',
                rule.security_profile or '',
                rule.category or '',
                rule.last_hit.strftime('%Y-%m-%d %H:%M:%S') if rule.last_hit else '',
                rule.description or ''
            ]
            for col, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col)
                cell.value = value
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # 컬럼 너비 자동 조정 (최대 400 픽셀)
        for column in worksheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # 너비 계산 (1 글자당 약 7 픽셀로 계산)
            adjusted_width = min(max_length + 2, 57)  # 57 = 400픽셀 / 7픽셀
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # 테두리 스타일 설정
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(results) + 1):
            for cell in row:
                cell.border = thin_border

        # 엑셀 파일 저장
        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='policies.xlsx'
        )

    except Exception as e:
        current_app.logger.error(f"정책 내보내기 중 오류 발생: {str(e)}")
        return jsonify({
            'success': False,
            'error': '정책 내보내기 중 오류가 발생했습니다.'
        }) 